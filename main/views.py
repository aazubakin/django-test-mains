from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import render
from rest_framework.generics import ListAPIView
from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView

from api.serializers import BillsSerializer
from main import models, forms, utils
from openpyxl import load_workbook
from django.views.generic import ListView
from filters import BillFilter, BillsApiFilter

# Create your views here.

menu = [
    {'title': "Форма загрузки файла клиентов", 'url': 'home'},
    {'title': "Форма загрузки файла счетов", 'url': 'form-bill'},
    {'title': "Список клиентов", 'url': 'client'},
    {'title': "Список счетов", 'url': 'bill'}
]


def form_bill(request):

    if request.POST:
        form = forms.AddBill(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = load_workbook(filename=file.file)
            utils.import_bills(wb)
    else:
        form = forms.AddBill()
    form_bill = forms.AddBill()
    context = {'menu': menu, 'title': 'Форма загрузки счетов'}
    context['form_bill'] = form_bill
    return render(request, 'main/form-bills.html', context=context)


def home(request):
    if request.POST:
        form = forms.AddClients(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if '.xlsx' in file.name and 'client' in file.name:
                wb = load_workbook(filename=file.file)
                utils.import_org(wb)
            else:
                return HttpResponse('Wrong extenssion or file name')
    else:
        form = forms.AddClients()
    form_client = forms.AddClients()
    context = {'menu': menu, 'title': 'Форма загрузки клиентов'}
    context['form_client'] = form_client
    return render(request, 'main/index.html', context=context)


class ClientsView(ListView):
    model = models.Client
    template_name = 'main/clients.html'
    context_object_name = 'clients'

    def get_queryset(self):
        return models.Client.objects.all()\
            .annotate(org_count=Count('organization'), sum_org=Sum('organization__bill__sum_bill'))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Список клиентов'
        context['menu'] = menu
        return context


class BillsView(ListView):
    model = models.Bill
    template_name = 'main/bills.html'
    context_object_name = 'bills'

    def get_queryset(self):
       return models.Bill.objects.all().select_related('organization')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Список счетов'
        context['menu'] = menu
        context['filter'] = BillFilter(self.request.GET, queryset=self.get_queryset())
        return context


class BillsApiView(ListAPIView):

    serializer_class = BillsSerializer
    queryset = models.Bill.objects.all().prefetch_related('organization')
    filterset_class = BillsApiFilter


class ClientFileUploadView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, format=None):
        file = request.FILES['file']
        utils.handle_uploaded_file(file, 'clients')


class BillFileUpload(APIView):
    parser_classes = (MultiPartParser, )

    def post(self, request, format=None):
        file = request.FILES['file']
        utils.handle_uploaded_file(file, 'bills')
